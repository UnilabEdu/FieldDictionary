from random import sample, randint
from flask.cli import with_appcontext
import click
from openpyxl import load_workbook
from os import path
from csv import reader

from openpyxl.cell.rich_text import CellRichText

from src.config import Config
from src.extensions import db
from src.models import Term, TermCategory, Category, ConnectedTerm, User, EnglishSynonym


@click.command("init_db")
@with_appcontext
def init_db():
    click.echo("Database creation in progress")
    db.drop_all()
    db.create_all()
    click.echo("Database created!")


@click.command("populate_db")
@with_appcontext
def populate_db():
    click.echo("Creating categories")
    filepath = path.join(Config.BASE_DIRECTORY, "dargebi.csv")
    with open(path.join(path.dirname(__file__), filepath), encoding='utf-8') as csv:
        csv_reader = reader(csv, delimiter=",")
        next(csv_reader)  # Skip the headers row
        for row in csv_reader:
            new_category = Category(name=row[1].strip())
            if int(row[2]) != 0:
                new_category.parent_id = row[2]
            db.session.add(new_category)
        db.session.commit()


    filepath = path.join(Config.BASE_DIRECTORY, "IATGE2.xlsx")
    workbook = load_workbook(filepath, rich_text=True)
    worksheet = workbook.worksheets[0]

    existing_categories = {category.name: category.id for category in Category.query.all()}

    for row in worksheet.iter_rows(min_row=2):
        def rich_to_html(row):
            text = ""
            if isinstance(row.value, CellRichText):
                hyperlink = row.hyperlink.target if row.hyperlink is not None else "#"
                for text_block in row.value:
                    text_value = text_block.text
                    if text_block.font.italic:
                        text_value = f"<i>{text_value}</i>"

                    if text_block.font.bold:
                        text_value = f"<b>{text_value}</b>"

                    if text_block.font.color.value == "FF0070C0":
                        text_value = f'<a href="{hyperlink}" target="_blank">{text_value}</a>'
                    else:
                        text_value = text_value.replace("-", "\n", 1)
                    text += text_value
            else:
                text = row.value
            return text

        if "=" in row[3].value:
            word = row[3].value.replace("=", "").strip()
            existing_term = Term.query.filter(Term.eng_word == word).first()
            EnglishSynonym(eng_word=row[2].value.strip(), term_id=existing_term.id).create()
            continue

        term_source = rich_to_html(row[6])
        definition_source = rich_to_html(row[8])
        definition = rich_to_html(row[7])
        context_source = rich_to_html(row[10])
        context = rich_to_html(row[9])

        new_term = Term(geo_word=row[3].value, eng_word=row[2].value, grammar_form=row[4].value,
                        term_source=term_source, stylistic_label = row[5].value,
                        definition=definition, definition_source=definition_source,
                        context=context, context_source=context_source,
                        comment=row[14].value)
        new_term.create()

        georgian_synonyms = row[12].value
        if georgian_synonyms != None:
            synonyms = georgian_synonyms.replace("\n", "").split("`")
            for synonym in synonyms:
                synonym_term = Term.query.filter(Term.geo_word == synonym.strip()).first()
                ConnectedTerm(term1_id=new_term.id, term2_id=synonym_term.id, is_synonym=True).create()

        english_synonyms = row[11].value
        if english_synonyms != None:
            synonyms = english_synonyms.replace("\n", "").split("`")
            for synonym in synonyms:
                synonym_term = Term.query.filter(Term.eng_word == synonym.strip()).first()
                ConnectedTerm(term1_id=new_term.id, term2_id=synonym_term.id, is_synonym=True, is_english=True).create()

        related_words = row[13].value
        if related_words != None:
            relations = related_words.replace("\n", "").split("`")
            for relation in relations:
                related_term = Term.query.filter(Term.eng_word == relation.strip()).first()
                ConnectedTerm(term1_id=new_term.id, term2_id=related_term.id, is_synonym=False).create()

        categories = str(row[0].value).replace("\n", "")
        subcategories = str(row[1].value).replace("\n", "")
        for category in categories.split("`"):
            category = category.strip()
            if not category or category == "None":
                continue

            if category not in existing_categories:
                new_category = Category(name=category)
                new_category.create(commit=False, flush=True)
                existing_categories[category] = new_category.id

            # If subcategory column is empty
            if not subcategories or subcategories == "None":
                term_category = TermCategory(term_id=new_term.id, category_id=existing_categories[category])
                term_category.create(commit=False)

        children_subcategories = subcategories.split("|")
        for index, subcategory in enumerate(children_subcategories):
            subcategory = subcategory.strip()
            if "`" in subcategory or not subcategory or subcategory == "None":
                continue

            if subcategory not in existing_categories:
                if index == 0:
                    parent_id = existing_categories[str(row[0].value)]
                else:
                    parent_id = existing_categories[children_subcategories[index-1]]

                new_category = Category(name=subcategory, parent_id=parent_id)
                new_category.create(commit=False, flush=True)
                existing_categories[subcategory] = new_category.id

            # For child categories, only save LAST child to term
            if index == len(children_subcategories) - 1:
                term_category = TermCategory(term_id=new_term.id, category_id=existing_categories[subcategory])
                term_category.create(commit=False)

        related_subcategories = subcategories.replace("\n", "").split("`")
        for subcategory in related_subcategories:
            subcategory = subcategory.strip()
            if "|" in subcategory or not subcategory or subcategory == "None":
                continue

            if subcategory not in existing_categories:
                parent_id = existing_categories[str(row[0].value)]
                new_category = Category(name=subcategory, parent_id=parent_id)
                new_category.create(commit=False, flush=True)
                existing_categories[subcategory] = new_category.id

            # For multiple related categories, save ALL of them to term
            term_category = TermCategory(term_id=new_term.id, category_id=existing_categories[subcategory])
            term_category.create(commit=False)
    db.session.commit()

    click.echo("Creating admin user...")
    admin = User(username="admin", password="admin123", email="testuser@gmail.com")

    admin.create()
    click.echo("Created admin user!")

    click.echo("Database populated!")
